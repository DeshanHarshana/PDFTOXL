"""
Write extracted PDF data into an Excel file that follows the template
structure of *output-excel-example.xlsx*.

Public API
----------
write_excel(records, output_path, template_path=None)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    DATA_START_ROW,
    EXCEL_COLUMNS,
    NUMBER_FORMATS,
    REJECTION_LIMIT,
)

MAX_COL = max(EXCEL_COLUMNS.keys())  # 14 = column N


# ---------------------------------------------------------------------------
# Template header creation (rows 1-5)
# ---------------------------------------------------------------------------

def _write_headers(ws: Any) -> None:
    """Write the fixed header rows (1-5) that mirror the template."""
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # Row 1-2: merged title
    ws.merge_cells("A1:M2")
    ws["A1"] = "CYLINDER HEAD MEASUREMENT RECORD SUMMARY"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = center

    header_font = Font(bold=True, size=14)

    # Row 3: group headers
    ws.merge_cells("B3:G3")
    ws["B3"] = "Stem diameter (\u00d8D1) in  mm"
    ws["B3"].font = header_font
    ws["B3"].alignment = center

    ws.merge_cells("H3:K3")
    ws["H3"] = "Valve guide (\u00d8d1) in mm"
    ws["H3"].font = header_font
    ws["H3"].alignment = center

    ws.merge_cells("L3:M3")
    ws["L3"] = "Clearance"
    ws["L3"].font = header_font
    ws["L3"].alignment = center

    # Row 4: sub-group headers
    ws.merge_cells("B4:D4")
    ws["B4"] = "INLET VALVE - A"
    ws["B4"].font = header_font
    ws["B4"].alignment = center

    ws.merge_cells("E4:G4")
    ws["E4"] = "INLET VALVE - B"
    ws["E4"].font = header_font
    ws["E4"].alignment = center

    ws.merge_cells("L4:M4")
    ws["L4"] = "Calculated"
    ws["L4"].font = header_font
    ws["L4"].alignment = center

    # Row 5: column names with background colours
    col_names = {
        1: "Cyl Head_SN",
        2: "A-I", 3: "A-II", 4: "A-III",
        5: "B-I", 6: "B-II", 7: "B-III",
        8: "A-VG-I", 9: "A-VG-II",
        10: "B-VG-I", 11: "B-VG-II",
        12: "A-Clearance", 13: "B-Clearance",
    }

    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_orange = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_ash = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    col_fills: dict[int, PatternFill] = {
        2: fill_green, 3: fill_green, 4: fill_green,       # A-I … A-III
        5: fill_orange, 6: fill_orange, 7: fill_orange,    # B-I … B-III
        8: fill_blue, 9: fill_blue, 10: fill_blue, 11: fill_blue,  # VG cols
        12: fill_ash, 13: fill_ash,                         # Clearance cols
    }

    for col_num, name in col_names.items():
        cell = ws.cell(row=5, column=col_num, value=name)
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center" if col_num == 1 else "right",
            vertical="center",
        )
        if col_num in col_fills:
            cell.fill = col_fills[col_num]

    # Apply thin border to header row
    thin = Side(style="thin")
    for col_num in range(1, MAX_COL + 1):
        ws.cell(row=5, column=col_num).border = Border(
            bottom=thin, top=thin, left=thin, right=thin,
        )


# ---------------------------------------------------------------------------
# Column widths (approximate match to template)
# ---------------------------------------------------------------------------

_COL_WIDTHS: dict[str, float] = {
    "A": 20.2, "B": 11.0, "C": 16.5, "D": 17.9,
    "E": 11.0, "F": 19.3, "G": 25.0, "H": 11.0, "I": 17.9,
    "J": 15.9, "K": 14.3, "L": 14.5, "M": 15.5,
}


def _set_column_widths(ws: Any) -> None:
    for letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Data rows
# ---------------------------------------------------------------------------

def _write_data_row(ws: Any, row: int, record: dict[str, Any]) -> None:
    """Write one data row from a parsed-PDF record dict."""
    for col_num, key in EXCEL_COLUMNS.items():
        cell = ws.cell(row=row, column=col_num)
        value = record.get(key)

        if value is not None and value != "":
            try:
                cell.value = float(value)
            except (ValueError, TypeError):
                cell.value = value
        else:
            cell.value = None

        cell.number_format = NUMBER_FORMATS.get(col_num, "General")
        cell.alignment = Alignment(
            horizontal="center" if col_num == 1 else "right"
        )


# ---------------------------------------------------------------------------
# Footer rows (COUNTIF + rejection limit)
# ---------------------------------------------------------------------------

def _write_footer(ws: Any, last_data_row: int) -> None:
    """Write the COUNTIF summary and rejection limit below the data."""
    gap_row = last_data_row + 1
    limit_row = last_data_row + 2
    label_row = last_data_row + 3

    first = DATA_START_ROW
    l_range = f"L{first}:L{last_data_row}"
    m_range = f"M{first}:M{last_data_row}"

    ws.cell(row=gap_row, column=12).value = (
        f'=COUNTIF({l_range},">="&L{limit_row})'
    )
    ws.cell(row=gap_row, column=12).alignment = Alignment(horizontal="center")

    ws.cell(row=gap_row, column=13).value = (
        f'=COUNTIF({m_range},">="&L{limit_row})'
    )
    ws.cell(row=gap_row, column=13).alignment = Alignment(horizontal="center")

    # Rejection limit value
    ws.merge_cells(
        start_row=limit_row, start_column=12,
        end_row=limit_row, end_column=13,
    )
    cell_limit = ws.cell(row=limit_row, column=12)
    cell_limit.value = REJECTION_LIMIT
    cell_limit.number_format = "0.000"
    cell_limit.alignment = Alignment(horizontal="center")

    # Label
    ws.merge_cells(
        start_row=label_row, start_column=12,
        end_row=label_row, end_column=13,
    )
    cell_label = ws.cell(row=label_row, column=12)
    cell_label.value = "Rejection Limit"
    cell_label.font = Font(bold=True)
    cell_label.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_excel(
    records: list[dict[str, Any]],
    output_path: str | Path,
    progress_callback: Any | None = None,
) -> Path:
    """Create the output Excel file from a list of parsed-PDF record dicts.

    Parameters
    ----------
    records : list[dict]
        Each dict has keys matching ``config.EXCEL_COLUMNS`` values.
    output_path : str | Path
        Where to save the ``.xlsx`` file.
    progress_callback : callable(current, total) | None
        Optional callback invoked after each row is written.

    Returns
    -------
    Path to the written file.
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_headers(ws)
    _set_column_widths(ws)

    total = len(records)
    for idx, record in enumerate(records):
        row = DATA_START_ROW + idx
        _write_data_row(ws, row, record)
        if progress_callback:
            progress_callback(idx + 1, total)

    if records:
        _write_footer(ws, DATA_START_ROW + total - 1)

    wb.save(str(output_path))
    wb.close()
    return output_path
